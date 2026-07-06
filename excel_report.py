"""계약된 최종 산출물: CTS 사전적격심사표 엑셀 생성.

criteria_seed12.jsonl 의 열 정의(SEED1_2 시트, column_order, allowed_verdicts)를
그대로 따르고, VLM 심사 결과 + 교차검증 + 참조 DB(여행경보)를 판정에 반영한다.

판정 철학: 명백한 근거가 있는 것만 Pass/Fail 자동판정, 애매하면 확인필요.
모든 판정에 근거를 남긴다 (셀 메모 + 판정근거 시트).

usage: python excel_report.py <screening_results.json> <cross_check.json> <out.xlsx>
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent

FILL = {
    "Pass": PatternFill("solid", fgColor="C6EFCE"),
    "YES": PatternFill("solid", fgColor="C6EFCE"),
    "해당": PatternFill("solid", fgColor="C6EFCE"),
    "Fail": PatternFill("solid", fgColor="FFC7CE"),
    "NO": PatternFill("solid", fgColor="FFC7CE"),
    "확인필요": PatternFill("solid", fgColor="FFEB9C"),
    "확인필요(외부)": PatternFill("solid", fgColor="BDD7EE"),
}

COUNTRY_ALIAS = {"키르기즈공화국": "키르기스스탄"}


# ── 제출건별 문서 인덱스 ─────────────────────────────────────────────
class Submission:
    def __init__(self, name: str):
        self.name = name
        self.country = name.split("_")[0]
        self.company = name.split("_")[1] if "_" in name else name
        self.seed = name.split("_")[-1] if "Seed" in name else "?"
        self.docs = []       # {file, doc_type, verdict, notes, signer, extracted, completeness}
        self.xcheck = {}

    def find(self, *keywords, in_filename=True, in_doctype=True):
        hits = []
        for d in self.docs:
            hay = ((d["file"] if in_filename else "") + " "
                   + (str(d["doc_type"]) if in_doctype else ""))
            if any(k in hay for k in keywords):
                hits.append(d)
        return hits

    def best_company_name(self):
        for d in self.find("사업자등록"):
            ex = d.get("extracted") or {}
            if ex.get("company_name"):
                return ex["company_name"]
        return self.company


def load_submissions(screening_path: Path, xcheck_path: Path) -> list[Submission]:
    records = json.loads(screening_path.read_text(encoding="utf-8"))
    xchecks = {x["submission"]: x for x in json.loads(xcheck_path.read_text(encoding="utf-8"))}
    subs: dict[str, Submission] = {}
    for r in records:
        s = subs.setdefault(r["submission"], Submission(r["submission"]))
        for d in (r.get("result") or {}).get("documents", []):
            if not isinstance(d, dict):
                continue
            s.docs.append({
                "file": r["file"], "doc_type": str(d.get("doc_type") or ""),
                "verdict": d.get("verdict"), "notes": d.get("notes") or "",
                "signer": d.get("signature_or_seal") or {},
                "extracted": d.get("extracted") or {},
                "completeness": d.get("completeness") or {},
            })
    for name, s in subs.items():
        s.xcheck = xchecks.get(name, {})
    return [subs[k] for k in sorted(subs)]


# ── 항목별 판정 로직 ────────────────────────────────────────────────
def doc_exists_verdict(s: Submission, *keywords, label: str):
    """서류 존재+상태 기반 판정: 없으면 Fail, 빈 양식이면 Fail, 이슈 있으면 확인필요."""
    hits = s.find(*keywords)
    if not hits:
        return "Fail", f"{label} 미제출 (파일·문서 미확인)"
    worst = min(hits, key=lambda d: {"fail": 0, "uncertain": 1, "pass": 2}.get(d["verdict"], 1))
    if worst["verdict"] == "fail":
        return "Fail", f"{worst['file']}: {worst['notes'] or '빈 양식/미작성'}"
    if worst["verdict"] == "uncertain":
        return "확인필요", f"{worst['file']}: {worst['notes'] or '판독 불확실'}"
    note = "; ".join(filter(None, (d["notes"] for d in hits)))[:200]
    return "Pass", f"{hits[0]['file']} 확인" + (f" ({note})" if note else "")


def judge(s: Submission, c: dict, refs: dict):
    """(판정값, 근거) — allowed_verdicts 를 준수한다."""
    cid, allowed = c["criterion_id"], c["allowed_verdicts"]

    if cid in ("S12-INF-01", "S12-INF-02"):
        return "확인필요", "실무자 직접 입력 항목 (심사표 규정)"

    if cid == "S12-CTY-01":  # 공모대상국 (외부: KOICA 목록)
        if not refs.get("eligible_countries"):
            return "확인필요", "KOICA 대상국 목록 미내장 — 안내서 33p 대조 필요"
        ok = s.country in refs["eligible_countries"]
        return ("YES" if ok else "NO"), f"{s.country}: KOICA 대상국 목록 대조"

    if cid == "S12-CTY-02":  # 여행금지 국가
        levels = refs.get("travel_advisory_levels", {}).get(s.country, [])
        return "확인필요", (f"{s.country} 여행경보: {', '.join(levels) or '경보 없음'}"
                        + (" — 지역단위 경보 존재, 사업지역 대조 필요" if levels else ""))

    if cid == "S12-CTY-03":  # 지역별 해당여부
        levels = refs.get("travel_advisory_levels", {}).get(s.country, [])
        bad = [x for x in levels if x in ("여행금지", "출국권고", "특별여행주의보")]
        if bad:
            return "확인필요", f"{s.country}에 {', '.join(bad)} 지역 존재 — 사업지역이 해당 지역인지 확인 필요"
        return "Pass", f"{s.country}: 3단계 이상 경보 지역 없음 (내장 DB 기준)"

    if cid == "S12-DOC-01":
        return doc_exists_verdict(s, "개요서", label="국문사업개요서")
    if cid == "S12-DOC-02":
        v, why = doc_exists_verdict(s, "제안서", label="사업제안서")
        return ("Fail" if v == "Fail" else v), why
    if cid == "S12-DOC-03":
        return "확인필요", "제안서 본문의 활동 주체(단일법인/컨소시엄) 사람 확인 필요"
    if cid == "S12-DOC-04":
        return doc_exists_verdict(s, "예산", label="예산계획서")

    if cid == "S12-DSQ-01":  # 결격사유 서약서 (날인 포함)
        hits = s.find("서약서")
        if not hits:
            return "Fail", "서약서 미제출"
        unsigned = [d for d in hits if d["verdict"] == "fail"]
        rep = next((c_ for c_ in s.xcheck.get("checks", [])
                    if c_.get("id") == "representative_match"
                    and c_.get("status") == "mismatch"), None)
        if unsigned:
            return "확인필요", f"서약서 날인 누락: {unsigned[0]['file']} — 보완요청 필요"
        if rep:
            return "확인필요", f"서명권자 불일치: {rep['detail'][:150]}"
        if any(d["verdict"] == "uncertain" for d in hits):
            u = next(d for d in hits if d["verdict"] == "uncertain")
            return "확인필요", f"{u['file']}: {u['notes'][:150]}"
        return "Pass", "서약서 제출 + 대표자 날인 확인"

    if cid == "S12-DSQ-02":
        return "확인필요", "KOICA 참여제한 여부 — 내부 시스템 확인 필요 (외부확인 큐)"

    if cid == "S12-ELG-01":  # 국내 법인 여부
        hits = s.find("사업자등록")
        if not hits:
            return "확인필요", "사업자등록증 미제출 — 국내 법인 여부 확인 불가"
        return "Pass", f"국세청 발급 사업자등록증 확인 ({hits[0]['file']})"

    if cid == "S12-ELG-02":
        return "확인필요", "사업자등록증 개업일 기준 업력 10년 이내 여부 — 개업일 확인 필요"
    if cid == "S12-ELG-03":
        return "확인필요", "재응모 횟수 — 내부 이력 확인 필요 (외부확인 큐)"

    if cid == "S12-ELG-04":  # 실적증빙
        v, why = doc_exists_verdict(s, "실적", label="실적증빙")
        return v, why

    if cid == "S12-BON-01":
        return "확인필요", "Seed 0 최종평가 우수기업 여부 — 내부 확인 필요"
    if cid in ("S12-BON-02", "S12-BON-03"):
        kw = "기술마켓" if cid == "S12-BON-02" else "혁신제품"
        hits = s.find(kw)
        if hits:
            return "확인필요", f"관련 증빙 추정 파일 존재({hits[0]['file']}) — 인증 유효성 확인 필요"
        return "증빙 미제출", f"{kw} 인증 증빙 미확인 (가점 미적용)"

    MIS = {
        "S12-MIS-01": (("공문",), "공문"),
        "S12-MIS-02": (("실적",), "실적 증빙 자료"),
        "S12-MIS-03": (("기타 증빙", "기타증빙", "기타 서류"), "기타 증빙 자료"),
        "S12-MIS-04": (("참여확인서", "참여인력"), "참여인력 참여확인서"),
        "S12-MIS-05": (("이력서",), "대표자 이력서"),
        "S12-MIS-06": (("개인정보", "동의서"), "개인정보수집이용동의서"),
        "S12-MIS-07": (("건강보험",), "건강보험자격득실확인서"),
        "S12-MIS-08": (("사업자등록",), "사업자등록증"),
        "S12-MIS-09": (("등기",), "법인등기사항전부증명서"),
        "S12-MIS-10": (("타기관", "지원금 내역"), "타기관 지원금 내역서"),
    }
    if cid in MIS:
        keywords, label = MIS[cid]
        v, why = doc_exists_verdict(s, *keywords, label=label)
        if cid == "S12-MIS-10" and v == "Fail":
            return "확인필요", "타기관 지원금 내역서 미확인 — 제출 대상 여부 포함 확인 필요"
        return v, why

    return "확인필요", "판정 로직 미구현 항목 — 사람 확인"


def clamp(verdict: str, allowed: list[str]) -> str:
    if verdict in allowed:
        return verdict
    return "확인필요" if "확인필요" in allowed else allowed[-1]


# ── 엑셀 생성 ───────────────────────────────────────────────────────
def build(screening_path: Path, xcheck_path: Path, out_path: Path):
    criteria = sorted((json.loads(l) for l in
                       open(BASE / "criteria/criteria_seed12.jsonl", encoding="utf-8")),
                      key=lambda c: c["column_order"])
    refs = json.loads((BASE / "data/reference_data.json").read_text(encoding="utf-8"))
    subs = load_submissions(screening_path, xcheck_path)

    wb = Workbook()
    head_font = Font(bold=True, size=9)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ① 심사표 시트 (계약 형식: SEED1_2)
    ws = wb.active
    ws.title = "SEED1_2"
    id_cols = ["순번", "국가", "기업명", "Seed"]
    for j, h in enumerate(id_cols, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font, cell.fill, cell.alignment = head_font, head_fill, wrap
    for c in criteria:
        cell = ws.cell(row=1, column=c["column_order"] - 5 + len(id_cols),
                       value=c["column_name"])
        cell.font, cell.fill, cell.alignment = head_font, head_fill, wrap
    ws.freeze_panes = "E2"

    detail_rows, remediation_rows = [], []
    for i, s in enumerate(subs, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=s.country)
        ws.cell(row=row, column=3, value=s.best_company_name())
        ws.cell(row=row, column=4, value=s.seed)
        for c in criteria:
            verdict, why = judge(s, c, refs)
            verdict = clamp(verdict, c["allowed_verdicts"])
            shown = ("확인필요(외부)"
                     if verdict == "확인필요" and c["judgment_mode"] == "external_check"
                     else verdict)
            col = c["column_order"] - 5 + len(id_cols)
            cell = ws.cell(row=row, column=col, value=shown)
            if shown in FILL:
                cell.fill = FILL[shown]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.comment = Comment(f"[{c['criterion_id']}] {why}", "CTS 자동심사", height=120, width=340)
            detail_rows.append((s.name, c["criterion_id"], c["item_name"],
                                c["severity"], shown, why))
            if shown in ("Fail", "확인필요") and c["severity"] in ("hard_fail", "remediable"):
                remediation_rows.append((s.name, c["item_name"], shown, why))

    for col, w in zip("ABCD", (5, 12, 22, 8)):
        ws.column_dimensions[col].width = w
    for j in range(len(id_cols) + 1, len(id_cols) + len(criteria) + 1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 14

    # ② 판정근거 시트
    ws2 = wb.create_sheet("판정근거")
    for j, h in enumerate(("제출건", "항목ID", "심사항목", "중요도", "판정", "근거"), start=1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font, cell.fill = head_font, head_fill
    for r, vals in enumerate(detail_rows, start=2):
        for j, v in enumerate(vals, start=1):
            ws2.cell(row=r, column=j, value=v).alignment = wrap
    for col, w in zip("ABCDEF", (28, 12, 30, 10, 12, 80)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ③ 보완요청 시트 (교차검증 요약 포함)
    ws3 = wb.create_sheet("보완요청")
    for j, h in enumerate(("제출건", "구분", "내용"), start=1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font, cell.fill = head_font, head_fill
    r = 2
    for s in subs:
        if s.xcheck.get("issues_summary"):
            ws3.cell(row=r, column=1, value=s.name)
            ws3.cell(row=r, column=2, value="교차검증 종합")
            ws3.cell(row=r, column=3, value=s.xcheck["issues_summary"]).alignment = wrap
            r += 1
        for miss in s.xcheck.get("missing_required_docs", []):
            ws3.cell(row=r, column=1, value=s.name)
            ws3.cell(row=r, column=2, value="필수서류 누락")
            ws3.cell(row=r, column=3, value=miss)
            r += 1
    for sub, item, verdict, why in remediation_rows:
        ws3.cell(row=r, column=1, value=sub)
        ws3.cell(row=r, column=2, value=f"{verdict}: {item}")
        ws3.cell(row=r, column=3, value=why).alignment = wrap
        r += 1
    for col, w in zip("ABC", (28, 24, 100)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    # ④ 종합 시트
    ws4 = wb.create_sheet("종합")
    for j, h in enumerate(("제출건", "국가", "기업명", "Pass", "Fail", "확인필요(내부)",
                           "확인필요(외부)", "hard_fail 이슈", "종합상태"), start=1):
        cell = ws4.cell(row=1, column=j, value=h)
        cell.font, cell.fill = head_font, head_fill
    by_sub = {}
    for sub, cid, item, sev, verdict, why in detail_rows:
        by_sub.setdefault(sub, []).append((sev, verdict))
    for r, s in enumerate(subs, start=2):
        vs = by_sub[s.name]
        n = lambda v: sum(1 for _, x in vs if x == v)
        hard = sum(1 for sev, x in vs if sev == "hard_fail" and x in ("Fail", "NO"))
        status = ("부적격 후보 (hard fail)" if hard
                  else "확인필요 잔여" if n("확인필요") + n("확인필요(외부)") else "적격 후보")
        for j, v in enumerate((s.name, s.country, s.best_company_name(), n("Pass"),
                               n("Fail"), n("확인필요"), n("확인필요(외부)"), hard, status), start=1):
            ws4.cell(row=r, column=j, value=v)
    for col, w in zip("ABCDEFGHI", (28, 12, 22, 7, 7, 13, 13, 13, 20)):
        ws4.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"저장: {out_path}")
    print(f"제출건 {len(subs)}건 × 항목 {len(criteria)}개 = 판정 {len(detail_rows)}건, "
          f"보완요청 {len(remediation_rows)}건")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    build(Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]))
